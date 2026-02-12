"""
审计日志接口
"""

from datetime import datetime
from typing import Optional
from io import BytesIO

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, desc, func, and_, or_

from core.database import get_db
from core.security import require_permission, require_manager, TokenData
from core.errors import BusinessException, ErrorCode
from schemas import paginate, success
from models import SystemLog, User

router = APIRouter(prefix="/api/v1/audit", tags=["审计"])


async def _build_query_conditions(
    level: Optional[str],
    module: Optional[str],
    action: Optional[str],
    start_time: Optional[datetime],
    end_time: Optional[datetime],
    keyword: Optional[str] = None,
    user_id: Optional[int] = None,
    username: Optional[str] = None
):
    """构建查询条件"""
    conditions = []
    if user_id:
        conditions.append(SystemLog.user_id == user_id)
    if username:
        conditions.append(or_(
            User.username.like(f"%{username}%"),
            User.nickname.like(f"%{username}%")
        ))
    if level and level != "undefined":
        conditions.append(SystemLog.level == level)
    if module and module != "undefined":
        conditions.append(SystemLog.module == module)
    if action and action != "undefined":
        conditions.append(SystemLog.action == action)
    if start_time:
        conditions.append(SystemLog.created_at >= start_time)
    if end_time:
        conditions.append(SystemLog.created_at <= end_time)
    
    if keyword:
        # 搜索消息内容、IP 地址或用户名/昵称
        conditions.append(or_(
            SystemLog.message.like(f"%{keyword}%"),
            SystemLog.ip_address.like(f"%{keyword}%"),
            User.username.like(f"%{keyword}%"),
            User.nickname.like(f"%{keyword}%")
        ))
        
    return and_(*conditions) if conditions else None


@router.get("")
async def list_logs(
    level: str = Query(None, description="INFO/WARNING/ERROR"),
    module: str = Query(None),
    action: str = Query(None),
    start_time: datetime = Query(None, description="开始时间"),
    end_time: datetime = Query(None, description="结束时间"),
    keyword: str = Query(None, description="关键词搜索(消息/IP)"),
    user_id: int = Query(None, description="用户ID筛选"),
    username: str = Query(None, description="用户名筛选"),
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=200),
    db: AsyncSession = Depends(get_db),
    user: TokenData = Depends(require_manager())
):
    """分页查询审计日志"""
    where_clause = await _build_query_conditions(level, module, action, start_time, end_time, keyword, user_id, username)

    # 统计总数 - 需要关联用户表以支持按用户名筛选
    count_stmt = (
        select(func.count())
        .select_from(SystemLog)
        .outerjoin(User, SystemLog.user_id == User.id)
    )
    if where_clause is not None:
        count_stmt = count_stmt.where(where_clause)
    total = (await db.execute(count_stmt)).scalar() or 0

    # 分页数据 - 使用 LEFT JOIN 关联用户表获取用户名和昵称
    data_stmt = (
        select(SystemLog, User.username, User.nickname)
        .outerjoin(User, SystemLog.user_id == User.id)
        .order_by(desc(SystemLog.created_at))
        .offset((page - 1) * size)
        .limit(size)
    )
    if where_clause is not None:
        data_stmt = data_stmt.where(where_clause)

    result = await db.execute(data_stmt)
    items = []
    for row in result.all():
        display_name = row.nickname or row.username
        if not display_name and row.SystemLog.user_id:
            display_name = f"用户#{row.SystemLog.user_id}"
        elif not display_name:
            display_name = "系统"
            
        items.append({
            "id": row.SystemLog.id,
            "level": row.SystemLog.level,
            "module": row.SystemLog.module,
            "action": row.SystemLog.action,
            "message": row.SystemLog.message,
            "username": display_name,
            "userId": row.SystemLog.user_id,
            "ip_address": row.SystemLog.ip_address,
            "user_agent": getattr(row.SystemLog, 'user_agent', None),
            "request_method": getattr(row.SystemLog, 'request_method', None),
            "request_path": getattr(row.SystemLog, 'request_path', None),
            "created_at": row.SystemLog.created_at
        })

    return paginate(items, total, page, size)


@router.get("/export")
async def export_logs(
    level: str = Query(None, description="INFO/WARNING/ERROR"),
    module: str = Query(None),
    action: str = Query(None),
    start_time: datetime = Query(None, description="开始时间"),
    end_time: datetime = Query(None, description="结束时间"),
    keyword: str = Query(None, description="关键词搜索(消息/IP)"),
    user_id: int = Query(None, description="用户ID筛选"),
    username: str = Query(None, description="用户名筛选"),
    db: AsyncSession = Depends(get_db),
    user: TokenData = Depends(require_manager())
):
    """导出审计日志为 Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        from fastapi import HTTPException
        raise BusinessException(ErrorCode.SERVICE_UNAVAILABLE, "导出功能需要 openpyxl 库支持")

    where_clause = await _build_query_conditions(level, module, action, start_time, end_time, keyword, user_id, username)

    # 分批查询数据（每批1000条，最多10000条），避免一次性加载全部到内存
    batch_size = 1000
    max_rows = 10000
    rows = []
    
    for offset in range(0, max_rows, batch_size):
        data_stmt = (
            select(SystemLog, User.username, User.nickname)
            .outerjoin(User, SystemLog.user_id == User.id)
            .order_by(desc(SystemLog.created_at))
            .offset(offset)
            .limit(batch_size)
        )
        if where_clause is not None:
            data_stmt = data_stmt.where(where_clause)

        result = await db.execute(data_stmt)
        batch = result.all()
        if not batch:
            break
        rows.extend(batch)

    # 创建 Excel 工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "审计日志"

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    # 写入表头
    headers = ["ID", "级别", "模块", "动作", "用户", "IP地址", "消息", "时间"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # 写入数据
    for row_idx, row in enumerate(rows, 2):
        display_name = row.nickname or row.username
        if not display_name and row.SystemLog.user_id:
            display_name = f"用户#{row.SystemLog.user_id}"
        elif not display_name:
            display_name = "系统"
            
        ws.cell(row=row_idx, column=1, value=row.SystemLog.id)
        ws.cell(row=row_idx, column=2, value=row.SystemLog.level)
        ws.cell(row=row_idx, column=3, value=row.SystemLog.module or "")
        ws.cell(row=row_idx, column=4, value=row.SystemLog.action or "")
        ws.cell(row=row_idx, column=5, value=display_name)
        ws.cell(row=row_idx, column=6, value=row.SystemLog.ip_address or "")
        ws.cell(row=row_idx, column=7, value=row.SystemLog.message or "")
        ws.cell(row=row_idx, column=8, value=row.SystemLog.created_at.strftime("%Y-%m-%d %H:%M:%S") if row.SystemLog.created_at else "")

    # 调整列宽
    column_widths = [8, 10, 15, 20, 15, 15, 50, 20]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # 生成文件名
    filename = f"审计日志_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    from urllib.parse import quote
    encoded_filename = quote(filename)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
        }
    )
