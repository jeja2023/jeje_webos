"""
数据导入/导出工具
支持 CSV、JSON、Excel 格式的数据导入导出
"""

import csv
import json
import logging
from typing import List, Dict, Any, Optional
from io import StringIO, BytesIO
from datetime import datetime

logger = logging.getLogger(__name__)

# 尝试导入 openpyxl（用于 Excel 支持）
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False
    logger.warning("openpyxl 未安装，Excel 导出功能不可用。请运行: pip install openpyxl")


class DataExporter:
    """数据导出器"""
    
    @staticmethod
    def export_to_csv(data: List[Dict[str, Any]], filename: Optional[str] = None) -> BytesIO:
        """
        导出数据为 CSV 格式
        
        Args:
            data: 数据列表
            filename: 文件名（可选）
        
        Returns:
            BytesIO 对象
        """
        if not data:
            return BytesIO()
        
        output = StringIO()
        fieldnames = data[0].keys()
        
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(data)
        
        # 转换为字节流
        result = BytesIO()
        result.write(output.getvalue().encode('utf-8-sig'))  # 使用 BOM 以支持 Excel
        result.seek(0)
        
        return result
    
    @staticmethod
    def export_to_json(data: List[Dict[str, Any]], indent: int = 2) -> str:
        """
        导出数据为 JSON 格式
        
        Args:
            data: 数据列表
            indent: JSON 缩进
        
        Returns:
            JSON 字符串
        """
        return json.dumps(data, ensure_ascii=False, indent=indent, default=str)
    
    @staticmethod
    def export_to_excel(data: List[Dict[str, Any]], sheet_name: str = "数据") -> BytesIO:
        """
        导出数据为 Excel 格式 (.xlsx)
        
        Args:
            data: 数据列表
            sheet_name: 工作表名称
        
        Returns:
            BytesIO 对象
        """
        if not EXCEL_SUPPORT:
            logger.warning("openpyxl 未安装，使用 CSV 替代")
            return DataExporter.export_to_csv(data)
        
        if not data:
            # 创建空的 Excel 文件
            wb = Workbook()
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output
        
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # 表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 写入表头
        headers = list(data[0].keys())
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 写入数据
        for row_idx, item in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = item.get(header, "")
                # 处理特殊类型
                if isinstance(value, datetime):
                    value = value.strftime("%Y-%m-%d %H:%M:%S")
                elif isinstance(value, (list, dict)):
                    value = json.dumps(value, ensure_ascii=False)
                
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")
        
        # 自动调整列宽
        for col_idx, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_idx)
            # 计算最大宽度
            max_length = len(str(header))
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        cell_length = len(str(cell.value)) if cell.value else 0
                        if cell_length > max_length:
                            max_length = cell_length
                    except Exception as e:
                        # 单元格计算失败时跳过，不影响整体导出
                        logger.debug(f"计算单元格宽度失败（已跳过）: {e}")
            adjusted_width = min(max_length + 2, 50)  # 最大宽度 50
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 冻结首行
        ws.freeze_panes = "A2"
        
        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output


class DataImporter:
    """数据导入器"""
    
    @staticmethod
    def import_from_csv(content: str) -> List[Dict[str, Any]]:
        """
        从 CSV 导入数据
        
        Args:
            content: CSV 内容字符串
        
        Returns:
            数据列表
        """
        try:
            reader = csv.DictReader(StringIO(content))
            return list(reader)
        except Exception as e:
            logger.error(f"CSV 导入失败: {e}")
            raise ValueError(f"CSV 格式错误: {str(e)}")
    
    @staticmethod
    def import_from_json(content: str) -> List[Dict[str, Any]]:
        """
        从 JSON 导入数据
        
        Args:
            content: JSON 内容字符串
        
        Returns:
            数据列表
        """
        try:
            data = json.loads(content)
            if isinstance(data, list):
                return data
            elif isinstance(data, dict):
                return [data]
            else:
                raise ValueError("JSON 格式错误：必须是数组或对象")
        except json.JSONDecodeError as e:
            logger.error(f"JSON 导入失败: {e}")
            raise ValueError(f"JSON 格式错误: {str(e)}")
    
    @staticmethod
    def import_from_excel(content: bytes) -> List[Dict[str, Any]]:
        """
        从 Excel 导入数据
        
        Args:
            content: Excel 文件内容（字节）
        
        Returns:
            数据列表
        """
        if not EXCEL_SUPPORT:
            raise ValueError("Excel 导入不可用：请安装 openpyxl (pip install openpyxl)")
        
        try:
            wb = load_workbook(filename=BytesIO(content), read_only=True)
            ws = wb.active
            
            # 获取表头
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)
            
            # 读取数据
            data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if all(cell is None for cell in row):
                    continue  # 跳过空行
                item = {}
                for idx, value in enumerate(row):
                    if idx < len(headers) and headers[idx]:
                        # 处理日期时间
                        if isinstance(value, datetime):
                            value = value.isoformat()
                        item[headers[idx]] = value
                data.append(item)
            
            return data
        except Exception as e:
            logger.error(f"Excel 导入失败: {e}")
            raise ValueError(f"Excel 格式错误: {str(e)}")
    
    @staticmethod
    def validate_data(data: List[Dict[str, Any]], required_fields: List[str]) -> tuple[bool, Optional[str]]:
        """
        验证导入数据
        
        Args:
            data: 数据列表
            required_fields: 必需字段列表
        
        Returns:
            (是否有效, 错误信息)
        """
        if not data:
            return False, "数据为空"
        
        for i, item in enumerate(data, 1):
            if not isinstance(item, dict):
                return False, f"第 {i} 行数据格式错误：必须是对象"
            
            for field in required_fields:
                if field not in item:
                    return False, f"第 {i} 行缺少必需字段: {field}"
        
        return True, None



